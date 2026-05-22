#coding:utf8
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

#获取当前路径
current_path = os.path.abspath(os.path.dirname(__file__))
#路径增加data文件夹
data_path = os.path.join(current_path, 'data')
original_file = os.path.join(data_path, 'history.xlsx')

# 根据给定文件和记录数据，输出excel文件
def output_excel(file, records,yymm):
    wb = Workbook()
    ws = wb.active
    ws.title = str(yymm)

    # 设置字体样式
    for row_idx, row_data in enumerate(records, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            #表头样式
            if row_idx == 1:
                cell.font = Font(name='微软雅黑', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                if col_idx == 7:
                    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                elif col_idx == 8:
                    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            else:
                # 数据行样式
                cell.font = Font(name='微软雅黑')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 4, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file)

# 根据年份读取历史记录文件中的original sheet中的one to six号码数据,返回一个元组，元组中的每个元素是每个号码的数组
def get_history_with_condition(file, year):
    df = pd.read_excel(file, sheet_name='original', dtype=str)
    year_num = df['年份'].astype(int)
    df = df[year_num < year]  # 筛选出年份小于指定年份的数据记录
    one_list = df['ONE'].tolist()
    two_list = df['TWO'].tolist()
    three_list = df['THREE'].tolist()
    four_list = df['FOUR'].tolist()
    five_list = df['FIVE'].tolist()
    six_list = df['SIX'].tolist()
    blue_list = df['蓝球'].tolist()

    return one_list, two_list, three_list, four_list, five_list, six_list, blue_list

# 根据条件读取原始数据读取历史记录文件中的original
history_one_list, history_two_list, history_three_list, history_four_list, history_five_list, history_six_list, history_blue_list = get_history_with_condition(original_file,2020)

header = ['期号','ONE','TWO','THREE','FOUR','FIVE','SIX','红球', '篮球', '年份','one_pos','two_pos','three_pos','four_pos','five_pos','six_pos','one_rate','two_rate','three_rate','four_rate','five_rate','six_rate','blue_rate','blue_pos']
records = [header]
#设定输出文件名
outputfile = os.path.join(data_path, f'trend.xlsx')

# 根据2023，2024，2025，2026 进行循环
for year in range(2020, 2027):
    # 根据条件读取原始数据读取历史记录文件中的original
    df = pd.read_excel(original_file, sheet_name='original', dtype=str)
    # 筛选出年份等于当前循环年份的数据记录，并对每条记录进行核对，核对内容包括每个号码出现的频率
    df = df[df['年份'] == str(year)]
    for index, row in df.iterrows():
        one = row['ONE']
        two = row['TWO']
        three = row['THREE']
        four = row['FOUR']
        five = row['FIVE']
        six = row['SIX']
        blue = row['蓝球']

        # 将历史记录中的号码数据转换为Series对象，方便后续统计频率
        s_one = pd.Series(history_one_list)
        s_two = pd.Series(history_two_list)
        s_three = pd.Series(history_three_list)
        s_four = pd.Series(history_four_list)
        s_five = pd.Series(history_five_list)
        s_six = pd.Series(history_six_list)
        s_blue = pd.Series(history_blue_list)

        s_one_rank = s_one.value_counts(normalize=True, ascending=True).round(4) # 统计one号码在历史记录中出现的频率
        s_two_rank = s_two.value_counts(normalize=True, ascending=True).round(4) # 统计two号码在历史记录中出现的频率
        s_three_rank = s_three.value_counts(normalize=True, ascending=True).round(4) # 统计three号码在历史记录中出现的频率
        s_four_rank = s_four.value_counts(normalize=True, ascending=True).round(4) # 统计four号码在历史记录中出现的频率
        s_five_rank = s_five.value_counts(normalize=True, ascending=True).round(4) # 统计five号码在历史记录中出现的频率
        s_six_rank = s_six.value_counts(normalize=True, ascending=True).round(4) # 统计six号码在历史记录中出现的频率
        s_blue_rank = s_blue.value_counts(normalize=True, ascending=True).round(4) # 统计blue号码在历史记录中出现的频率
        one_rate = s_one_rank.get(one, 0)  # 计算one号码在历史记录中出现的频率，保留4位小数
        two_rate = s_two_rank.get(two, 0)  # 计算two号码在历史记录中出现的频率，保留4位小数
        three_rate = s_three_rank.get(three, 0)  # 计算three号码在历史记录中出现的频率，保留4位小数
        four_rate = s_four_rank.get(four, 0)  # 计算four号码在历史记录中出现的频率，保留4位小数
        five_rate = s_five_rank.get(five, 0)  # 计算five号码在历史记录中出现的频率，保留4位小数
        six_rate = s_six_rank.get(six, 0)  # 计算six号码在历史记录中出现的频率，保留4位小数
        blue_rate = s_blue_rank.get(blue, 0)  # 计算blue号码在历史记录中出现的频率，保留4位小数

        one_pos = s_one_rank.index.get_loc(one) +1 if one in s_one_rank.index else 0  # 计算one号码在历史记录中出现的排名位置，排名从1开始，如果号码未出现过，则排名位置为0
        two_pos = s_two_rank.index.get_loc(two) +1 if two in s_two_rank.index else 0  # 计算two号码在历史记录中出现的排名位置，排名从1开始，如果号码未出现过，则排名位置为0
        three_pos = s_three_rank.index.get_loc(three) +1 if three in s_three_rank.index else 0  # 计算three号码在历史记录中出现的排名位置，排名从1开始，如果号码未出现过，则排名位置为0
        four_pos = s_four_rank.index.get_loc(four) +1 if four in s_four_rank.index else 0  # 计算four号码在历史记录中出现的排名位置，排名从1开始，如果号码未出现过，则排名位置为0
        five_pos = s_five_rank.index.get_loc(five) +1 if five in s_five_rank.index else 0  # 计算five号码在历史记录中出现的排名位置，排名从1开始，如果号码未出现过，则排名位置为0
        six_pos = s_six_rank.index.get_loc(six) +1 if six in s_six_rank.index else 0  # 计算six号码在历史记录中出现的排名位置，排名从1开始，如果号码未出现过，则排名位置为0
        blue_pos = s_blue_rank.index.get_loc(blue) +1 if blue in s_blue_rank.index else 0  # 计算blue号码在历史记录中出现的排名位置，排名从1开始，如果号码未出现过，则排名位置为0

        # 当前红码号码添加到历史记录中，供后续核对使用
        history_one_list.append(one)
        history_two_list.append(two)
        history_three_list.append(three)
        history_four_list.append(four)
        history_five_list.append(five)
        history_six_list.append(six)
        history_blue_list.append(blue)

        # 格式化输出数据记录
        data = row[0:10].tolist() + [one_pos,two_pos,three_pos,four_pos,five_pos,six_pos, one_rate,two_rate,three_rate,four_rate,five_rate,six_rate,blue_pos,blue_rate]
        records.append(data)
            

# 记录数据写入输出文件
output_excel(outputfile, records, year)