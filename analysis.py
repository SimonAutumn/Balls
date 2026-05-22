#coding:utf8
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

#获取当前路径
current_path = os.path.abspath(os.path.dirname(__file__))
#路径增加data文件夹
data_path = os.path.join(current_path, 'data')
original_file = os.path.join(data_path, 'history.xlsx')

# 根据给定文件和记录数据，输出excel文件
def output_excel(file, records,yymm):
    wb = Workbook()
    ws = wb.active
    ws.title = str(yymm)

    # 设置字体样式
    for row_idx, row_data in enumerate(records, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            #表头样式
            if row_idx == 1:
                cell.font = Font(name='微软雅黑', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                if col_idx == 7:
                    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                elif col_idx == 8:
                    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            else:
                # 数据行样式
                cell.font = Font(name='微软雅黑')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 6, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file)

# 根据年份读取历史记录文件中的original sheet中的红球、年份、前五、前四、后五、后四、前三、后三数据,返回红球、前五、前四、后五、后四、前三、后三数组
def get_history_with_condition(file, year):
    df = pd.read_excel(file, sheet_name='original', dtype=str)
    red_balls = []
    top5 = []
    top4 = []
    after5 = []
    after4 = []
    top3 = []
    after3 = []
    for index, row in df.iterrows():
        if int(row['年份']) < year:
            red_balls.append(row['红球'])
            top5.append(row['前五'])
            top4.append(row['前四'])
            after5.append(row['后五'])
            after4.append(row['后四'])
            top3.append(row['前三'])
            after3.append(row['后三'])
    return red_balls, top5, top4, after5, after4, top3, after3

# 根据条件读取原始数据读取历史记录文件中的original
history_red_balls, history_top5, history_top4, history_after5, history_after4, history_top3, history_after3 = get_history_with_condition(original_file,2020)

header = ['期号','ONE','TOW','THREE','FOUR','FIVE','SIX','红球', '篮球', '年份','重复标记', '前五标记', '前四标记', '后五标记', '后四标记', '前三标记', '后三标记']
records = [header]
#设定输出文件名
outputfile = os.path.join(data_path, f'check.xlsx')

# 根据2020，2021，2022，2023，2024，2025，2026 进行循环
for year in range(2020, 2027):
    # 根据条件读取原始数据读取历史记录文件中的original
    df = pd.read_excel(original_file, sheet_name='original', dtype=str)
    for index, row in df.iterrows():
        if row['年份'] == str(year):
            red_ball = str(row['红球'])
            top5 = row['前五']
            top4 = row['前四']
            after5 = row['后五']
            after4 = row['后四']
            top3 = row['前三']
            after3 = row['后三']

            # 核对红球号码是否在历史记录中出现过，出现过则标记为1，否则标记为0
            red_ball_mark = 1 if red_ball in history_red_balls else 0
            history_red_balls.append(red_ball)  # 将当前红球号码添加到历史记录中，供后续核对使用

            # 核对前五是否在历史记录中出现过，出现过则标记为1，否则标记为0
            top5_mark = 1 if top5 in history_top5 else 0
            history_top5.append(top5)  # 将当前前五数据添加到历史记录中，供后续核对使用

            # 核对前四是否在历史记录中出现过，出现过则标记为1，否则标记为0
            top4_mark = 1 if top4 in history_top4 else 0
            history_top4.append(top4)  # 将当前前四数据添加到历史记录中，供后续核对使用 

            # 核对后五是否在历史记录中出现过，出现过则标记为1，否则标记为0
            after5_mark = 1 if after5 in history_after5 else 0
            history_after5.append(after5)  # 将当前后五数据添加到历史记录中，供后续核对使用

            # 核对后四是否在历史记录中出现过，出现过则标记为1，否则标记为0
            after4_mark = 1 if after4 in history_after4 else 0
            history_after4.append(after4)  # 将当前后四数据添加到历史记录中，供后续核对使用

            # 核对前三是否在历史记录中出现过，出现过则标记为1，否则标记为0
            top3_mark = 1 if top3 in history_top3 else 0
            history_top3.append(top3)  # 将当前前三数据添加到历史记录中，供后续核对使用

            # 核对后三是否在历史记录中出现过，出现过则标记为1，否则标记为0
            after3_mark = 1 if after3 in history_after3 else 0
            history_after3.append(after3)  # 将当前后三数据添加到历史记录中，供后续核对使用

            # 格式化输出数据记录
            data = row[0:10].tolist() + [red_ball_mark, top5_mark, top4_mark, after5_mark, after4_mark, top3_mark, after3_mark]
            records.append(data)
            

# 记录数据写入输出文件
output_excel(outputfile, records, 2026)

