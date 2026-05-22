#coding:utf8
import os
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

#获取当前路径
current_path = os.path.abspath(os.path.dirname(__file__))
#路径增加data文件夹
data_path = os.path.join(current_path, 'data')
original_file = os.path.join(data_path, 'history.xlsx')

# 根据给定文件和记录数据，输出excel文件
def output_excel(file, records,yymm):
    wb = Workbook()

    # 判断文件是否存在，如果存在则加载文件，否则创建新文件
    if os.path.exists(file):
        wb = openpyxl.load_workbook(file)
        ws = wb.create_sheet(str(yymm))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = str(yymm)

    # 设置字体样式
    for row_idx, row_data in enumerate(records, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            #表头样式
            if row_idx == 1:
                cell.font = Font(name='微软雅黑', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                if col_idx == 7:
                    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                else:
                    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                # 数据行样式
                cell.font = Font(name='微软雅黑')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 4, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file)

# 读取历史记录文件中的original sheet中的one to six号码数据,返回一个元组，元组中的每个元素是每个号码站比率的数组
def get_history(file):
    df = pd.read_excel(file, sheet_name='original', dtype=str)
    one_list = df['ONE']
    two_list = df['TWO']
    three_list = df['THREE']
    four_list = df['FOUR']
    five_list = df['FIVE']
    six_list = df['SIX']
    blue_list = df['蓝球']

    return one_list, two_list, three_list, four_list, five_list, six_list, blue_list

# 根据条件读取原始数据读取历史记录文件中的original
history_one, history_two, history_three, history_four, history_five, history_six, history_blue= get_history(original_file)

#设定输出文件名
outputfile = os.path.join(data_path, f'test.xlsx')

s_one= history_one.value_counts(normalize=True, ascending=True).round(4) # 统计one号码在历史记录中出现的频率
s_two = history_two.value_counts(normalize=True, ascending=True).round(4) # 统计two号码在历史记录中出现的频率
s_three = history_three.value_counts(normalize=True, ascending=True).round(4) # 统计three号码在历史记录中出现的频率
s_four = history_four.value_counts(normalize=True, ascending=True).round(4) # 统计four号码在历史记录中出现的频率
s_five = history_five.value_counts(normalize=True, ascending=True).round(4) # 统计five号码在历史记录中出现的频率
s_six = history_six.value_counts(normalize=True, ascending=True).round(4) # 统计six号码在历史记录中出现的频率
s_blue = history_blue.value_counts(normalize=True, ascending=True).round(4) # 统计blue号码在历史记录中出现的频率

# 收集所有索引
index_series = [s_one.index.to_series(name='ONE').reset_index(drop=True), 
                s_two.index.to_series(name='TWO').reset_index(drop=True), 
                s_three.index.to_series(name='THREE').reset_index(drop=True),
                s_four.index.to_series(name='FOUR').reset_index(drop=True),
                s_five.index.to_series(name='FIVE').reset_index(drop=True),
                s_six.index.to_series(name='SIX').reset_index(drop=True),
                s_blue.index.to_series(name='BLUE').reset_index(drop=True)
                ]
# 按行合并（concat axis=1 会自动对齐不同的行索引）
df = pd.concat(index_series, axis=1)

# 获取当期日期的月日
from datetime import datetime
today = datetime.today()
yymm = today.strftime('%m%d')
# 记录数据写入输出文件
records = [df.columns.tolist()] + df.fillna('').values.tolist()
output_excel(outputfile, records, yymm)